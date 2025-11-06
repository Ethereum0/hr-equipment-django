# check_excel_structure.py
from openpyxl import load_workbook

def check_structure():
    wb = load_workbook('ТМЦ макет.xlsx', data_only=True)
    
    print("=== ПРОВЕРКА СТРУКТУРЫ EXCEL ===")
    
    # Проверяем вкладку 'Выдано'
    if 'Выдано' in wb.sheetnames:
        ws = wb['Выдано']
        print("\n--- ВКЛАДКА 'ВЫДАНО' ---")
        for i in range(1, 20):  # Смотрим первые 20 строк
            row = next(ws.iter_rows(min_row=i, max_row=i, values_only=True))
            print(f"Строка {i}: {row}")
            if i >= 3 and 'ФИО' in str(row[0]):  # Нашли заголовки
                print(f"✅ ЗАГОЛОВКИ НА СТРОКЕ {i}")
                break
    
    # Проверяем вкладку 'Employment' 
    if 'Employment' in wb.sheetnames:
        ws = wb['Employment']
        print("\n--- ВКЛАДКА 'EMPLOYMENT' ---")
        for i in range(1, 10):
            row = next(ws.iter_rows(min_row=i, max_row=i, values_only=True))
            print(f"Строка {i}: {row[0:5]}...")  # Первые 5 колонок
    
    # Проверяем вкладку 'Dismissal'
    if 'Dismissal' in wb.sheetnames:
        ws = wb['Dismissal']
        print("\n--- ВКЛАДКА 'DISMISSAL' ---")
        for i in range(1, 10):
            row = next(ws.iter_rows(min_row=i, max_row=i, values_only=True))
            print(f"Строка {i}: {row[0:5]}...")

if __name__ == "__main__":
    check_structure()