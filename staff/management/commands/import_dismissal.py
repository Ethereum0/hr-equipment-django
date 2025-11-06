import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee, Equipment

class Command(BaseCommand):
    help = 'Импорт уволенных/декретных из вкладки Dismissal'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)

        if 'Dismissal' in wb.sheetnames:
            ws = wb['Dismissal']
            self.parse_dismissal(ws)

        self.stdout.write(self.style.SUCCESS('✅ Импорт Dismissal завершён!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def parse_dismissal(self, ws):
        self.stdout.write("📊 Начало парсинга листа 'Dismissal'")
        
        # Заголовки в строке 1
        headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers_row)]
        
        col_map = {}
        for i, h in enumerate(headers):
            if 'ФИ' in h and i + 1 < len(headers) and 'О' in headers[i + 1]:
                col_map['first_last'] = i
                col_map['middle'] = i + 1
            elif 'AD' in h:
                col_map['ad_login'] = i

        updated_count = 0
        equipment_freed = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Пропускаем пустые строки
                continue
                
            first_last = self.clean_cell(row[col_map.get('first_last', 1)]) if col_map.get('first_last') is not None else ''
            middle = self.clean_cell(row[col_map.get('middle', 2)]) if col_map.get('middle') is not None else ''
            ad_login = self.clean_cell(row[col_map.get('ad_login', 3)]) if col_map.get('ad_login') is not None else ''
            
            # Формируем ФИО
            if first_last and middle:
                fio = f"{first_last} {middle}"
            else:
                fio = first_last
                
            if not fio:
                continue
                
            # Определяем статус (декрет или уволен)
            status = 'maternity' if 'декрет' in fio.lower() else 'dismissed'
            
            # Ищем сотрудника
            employee = None
            if ad_login:
                employee = Employee.objects.filter(ad_login=ad_login).first()
                
            if not employee:
                # Ищем по ФИО
                employee = Employee.objects.filter(fio__icontains=fio).first()
                
            if employee:
                # Обновляем статус
                if employee.status != status:
                    old_status = employee.status
                    employee.status = status
                    employee.save()
                    updated_count += 1
                    self.stdout.write(self.style.SUCCESS(f"✅ Обновлен статус: {employee.fio} -> {status} (было: {old_status})"))
                    
                    # Освобождаем оборудование
                    freed_count = Equipment.objects.filter(employee=employee).update(employee=None)
                    if freed_count > 0:
                        equipment_freed += freed_count
                        self.stdout.write(self.style.SUCCESS(f"✅ Освобождено оборудование: {freed_count} единиц у {employee.fio}"))
            else:
                self.stdout.write(self.style.WARNING(f"⚠️ Не найден в базе: {fio}"))
                
        self.stdout.write(self.style.SUCCESS(f"📊 Итоги: обновлено {updated_count} сотрудников, освобождено {equipment_freed} оборудования"))