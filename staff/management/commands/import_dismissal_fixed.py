# staff/management/commands/import_dismissal_fixed.py
import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee, Equipment

class Command(BaseCommand):
    help = 'Импорт уволенных/декретных из вкладки Dismissal с созданием отсутствующих сотрудников'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)

        if 'Dismissal' in wb.sheetnames:
            ws = wb['Dismissal']
            self.parse_dismissal(ws)

        self.stdout.write(self.style.SUCCESS('✅ Импорт Dismissal завершён!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def parse_dismissal(self, ws):
        self.stdout.write("📊 Начало парсинга листа 'Dismissal'")
        
        updated_count = 0
        created_count = 0
        equipment_freed = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[1]:  # Пропускаем пустые строки (колонка ФИ)
                    continue
                    
                first_last = self.clean_cell(row[1])  # ФИ
                middle = self.clean_cell(row[2])      # Отчество
                ad_login = self.clean_cell(row[3])    # AD логин
                email = self.clean_cell(row[5]) if len(row) > 5 else ''  # Email
                sip_number = self.clean_cell(row[7]) if len(row) > 7 else ''  # SIP номер
                phone = self.clean_cell(row[12]) if len(row) > 12 else ''  # телефон
                
                # Формируем ФИО
                if first_last and middle:
                    fio = f"{first_last} {middle}"
                else:
                    fio = first_last
                    
                if not fio:
                    continue
                
                # Разбиваем ФИО на компоненты
                parts = fio.split()
                last_name = parts[0] if len(parts) > 0 else ''
                first_name = parts[1] if len(parts) > 1 else ''
                
                # Определяем статус (декрет или уволен)
                status = 'maternity' if 'декрет' in fio.lower() else 'dismissed'
                
                # Ищем сотрудника
                employee = None
                if ad_login:
                    employee = Employee.objects.filter(ad_login=ad_login).first()
                    
                if not employee:
                    # Ищем по ФИО
                    employee = Employee.objects.filter(fio__icontains=fio).first()
                    
                if employee:
                    # Обновляем статус существующего сотрудника
                    if employee.status != status:
                        old_status = employee.status
                        employee.status = status
                        employee.save()
                        updated_count += 1
                        self.stdout.write(self.style.SUCCESS(f"✅ Обновлен статус: {employee.fio} -> {status} (было: {old_status})"))
                        
                        # Освобождаем оборудование
                        freed_count = Equipment.objects.filter(employee=employee).update(employee=None)
                        if freed_count > 0:
                            equipment_freed += freed_count
                            self.stdout.write(self.style.SUCCESS(f"✅ Освобождено оборудование: {freed_count} единиц у {employee.fio}"))
                else:
                    # СОЗДАЕМ НОВОГО СОТРУДНИКА С СТАТУСОМ "УВОЛЕН" или "ДЕКРЕТ"
                    # Обрабатываем email (если это формула CONCATENATE)
                    if email and '=CONCATENATE' in email:
                        # Извлекаем email из формулы
                        if 'p-el.ru' in email and ad_login:
                            email = f"{ad_login}@p-el.ru"
                    
                    # Создаем сотрудника
                    employee_data = {
                        'fio': fio,
                        'first_name': first_name,
                        'last_name': last_name,
                        'middle_name': middle,
                        'ad_login': ad_login if ad_login else None,
                        'email': email if email else '',
                        'sip_number': str(sip_number) if sip_number else '',
                        'phone': phone if phone else '',
                        'office': 'remote',
                        'status': status,
                    }
                    
                    # Создаем сотрудника
                    Employee.objects.create(**employee_data)
                    created_count += 1
                    self.stdout.write(self.style.SUCCESS(f"➕ Создан {status}: {fio}"))
                    
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"❌ Ошибка в строке {row_num}: {str(e)}"))
                import traceback
                self.stdout.write(self.style.ERROR(f"   Детали: {traceback.format_exc()}"))
                
        self.stdout.write(self.style.SUCCESS(f"📊 Итоги:"))
        self.stdout.write(self.style.SUCCESS(f"   Обновлено статусов: {updated_count}"))
        self.stdout.write(self.style.SUCCESS(f"   Создано уволенных/декретных: {created_count}"))
        self.stdout.write(self.style.SUCCESS(f"   Освобождено оборудования: {equipment_freed}"))