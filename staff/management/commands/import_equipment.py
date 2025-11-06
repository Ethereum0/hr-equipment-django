import os
import logging
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee, Equipment

# Настройка логирования
logger = logging.getLogger('import_equipment')

class Command(BaseCommand):
    help = 'Импорт оборудования из вкладки Выдано'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)

        if 'Выдано' in wb.sheetnames:
            ws = wb['Выдано']
            self.parse_equipment(ws)
        else:
            self.stdout.write(self.style.WARNING("Лист 'Выдано' не найден в файле."))

        self.stdout.write(self.style.SUCCESS('✅ Импорт Выдано завершён!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def find_employee_by_name(self, name_from_excel):
        """Ищет сотрудника по имени из Excel (только фамилия и имя)"""
        if not name_from_excel:
            return None
            
        name_clean = name_from_excel.strip()
        
        # В Excel только "Фамилия Имя", в БД может быть "Фамилия Имя Отчество"
        # Ищем сотрудников, у которых ФИО начинается с этой пары
        employees = Employee.objects.filter(fio__startswith=name_clean)
        
        if employees.count() == 1:
            employee = employees.first()
            self.stdout.write(self.style.SUCCESS(f"✅ Найден сотрудник: {employee.fio}"))
            return employee
        elif employees.count() > 1:
            # Если несколько совпадений, берем первого
            employee = employees.first()
            self.stdout.write(self.style.WARNING(f"⚠️ Несколько совпадений для '{name_clean}', берем: {employee.fio}"))
            return employee
        else:
            # Если не нашли по началу, пробуем поиск по фамилии
            parts = name_clean.split()
            if len(parts) >= 1:
                last_name = parts[0]
                employee = Employee.objects.filter(last_name__icontains=last_name).first()
                if employee:
                    self.stdout.write(self.style.SUCCESS(f"✅ Найден по фамилии: {employee.fio}"))
                    return employee
            
            self.stdout.write(self.style.WARNING(f"❌ Сотрудник не найден: '{name_clean}'"))
            return None

    def parse_equipment(self, ws):
        self.stdout.write("📊 Начало парсинга листа 'Выдано'")
        
        # Заголовки в СТРОКЕ 1
        headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers_row)]

        self.stdout.write(f"📋 Заголовки: {headers[:10]}...")  # Показываем первые 10
        
        col_map = {}
        for i, h in enumerate(headers):
            # ИСПРАВЛЕНО: правильный поиск столбца
            if 'ФИО сотрудника' in h:
                col_map['fio'] = i
                self.stdout.write(f"✅ Найден столбец 'ФИО сотрудника' в позиции {i}")
            elif 'должность' in h.lower():
                col_map['position'] = i
            elif 'офис' in h.lower():
                col_map['office'] = i
            elif 'Перечень ТМЦ' in h:
                col_map['type'] = i
            elif 'Модель' in h:
                col_map['model'] = i
            elif 'Серийный номер' in h:
                col_map['serial'] = i
            elif 'MAC адрес' in h:
                col_map['mac'] = i
            elif 'Коментарий (IP/AnyDesk)' in h:
                col_map['ip_anydesk'] = i
            elif 'Коментарий' in h and i not in col_map.values():
                col_map['comment'] = i
            elif 'Ноут в домене' in h:
                col_map['domain_note'] = i

        # Проверка обязательных столбцов
        if 'fio' not in col_map:
            self.stdout.write(self.style.ERROR("❌ Столбец 'ФИО сотрудника' не найден!"))
            self.stdout.write(self.style.ERROR(f"📋 Найденные столбцы: {list(col_map.keys())}"))
            return
            
        if 'type' not in col_map:
            self.stdout.write(self.style.ERROR("❌ Столбец 'Перечень ТМЦ' не найден!"))
            return

        self.stdout.write(self.style.SUCCESS("✅ Все обязательные столбцы найдены"))
        
        processed = 0
        created = 0
        skipped = 0
        employee_not_found = 0
        free_equipment = 0
        
        # Данные начинаются со СТРОКИ 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_list = [self.clean_cell(cell) for cell in row]
            
            # Пропускаем полностью пустые строки
            if not any(row_list):
                skipped += 1
                continue
                
            fio_excel = row_list[col_map['fio']]
            if not fio_excel:
                skipped += 1
                continue

            # Находим сотрудника (в Excel только "Фамилия Имя")
            employee = self.find_employee_by_name(fio_excel)

            # Получаем данные оборудования
            type_name = row_list[col_map['type']]
            model = row_list[col_map['model']] if 'model' in col_map else ''
            serial = row_list[col_map['serial']] if 'serial' in col_map else ''
            mac = row_list[col_map['mac']] if 'mac' in col_map else ''
            ip_anydesk = row_list[col_map['ip_anydesk']] if 'ip_anydesk' in col_map else ''
            comment = row_list[col_map['comment']] if 'comment' in col_map else ''
            position = row_list[col_map['position']] if 'position' in col_map else ''
            office = row_list[col_map['office']] if 'office' in col_map else 'remote'
            
            # Нормализация офиса
            office_normalized = office.upper() if office else 'REMOTE'
            valid_offices = [choice[0] for choice in Employee.OFFICE_CHOICES]
            if office_normalized not in valid_offices:
                office_normalized = 'REMOTE'
            
            # Обновляем должность сотрудника, если сотрудник найден и должность указана
            if employee and position and (not employee.position or employee.position.strip() == ''):
                employee.position = position
                employee.save()
                self.stdout.write(self.style.SUCCESS(f"💼 Обновлена должность: {employee.fio} -> '{position}'"))
                
            # Создаем/обновляем оборудование
            if type_name and serial:
                equipment, created_flag = Equipment.objects.update_or_create(
                    serial_number=serial,
                    defaults={
                        'employee': employee,  # ✅ Может быть None - тогда оборудование свободное
                        'type': type_name,
                        'model': model,
                        'mac_address': mac,
                        'ip_or_anydesk': ip_anydesk,
                        'comment': comment,
                        'office': office_normalized,
                    }
                )
                
                if created_flag:
                    created += 1
                    if employee:
                        self.stdout.write(self.style.SUCCESS(f"➕ Оборудование: {type_name} ({serial}) для {employee.fio}"))
                    else:
                        free_equipment += 1
                        self.stdout.write(self.style.WARNING(f"📦 Свободное оборудование: {type_name} ({serial}) - сотрудник '{fio_excel}' не найден"))
                else:
                    if employee:
                        self.stdout.write(self.style.WARNING(f"🔄 Обновлено оборудование: {type_name} ({serial}) для {employee.fio}"))
                    else:
                        self.stdout.write(self.style.WARNING(f"🔄 Обновлено свободное оборудование: {type_name} ({serial})"))
                    
            processed += 1
            
            if not employee:
                employee_not_found += 1
            
        self.stdout.write(self.style.SUCCESS(f"📊 Итоги импорта:"))
        self.stdout.write(self.style.SUCCESS(f"   Обработано строк: {processed}"))
        self.stdout.write(self.style.SUCCESS(f"   Создано оборудования: {created}"))
        self.stdout.write(self.style.SUCCESS(f"   Свободное оборудование: {free_equipment}"))
        self.stdout.write(self.style.WARNING(f"   Пропущено пустых: {skipped}"))
        self.stdout.write(self.style.WARNING(f"   Сотрудники не найдены: {employee_not_found}"))