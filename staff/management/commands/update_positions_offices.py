# staff/management/commands/update_positions_offices.py
import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee

class Command(BaseCommand):
    help = 'Обновление должностей и офисов из вкладки Выдано'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)
        
        if 'Выдано' in wb.sheetnames:
            ws = wb['Выдано']
            self.update_from_equipment(ws)
            
        self.stdout.write(self.style.SUCCESS('✅ Обновление должностей и офисов завершено!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def update_from_equipment(self, ws):
        self.stdout.write("📊 Обновление из вкладки Выдано")
        
        updated_positions = 0
        updated_offices = 0
        
        # Пропускаем заголовок
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:  # Пропускаем пустые строки (ФИО сотрудника)
                    continue
                    
                fio_excel = self.clean_cell(row[0])  # ФИО сотрудника
                position = self.clean_cell(row[1])   # должность
                office = self.clean_cell(row[2])     # офис
                
                if not fio_excel:
                    continue
                
                # Ищем сотрудника по ФИО (в Excel только "Фамилия Имя")
                employee = Employee.objects.filter(fio__startswith=fio_excel).first()
                
                if employee:
                    update_fields = []
                    
                    # Обновляем должность
                    if position and (not employee.position or employee.position.strip() == ''):
                        employee.position = position
                        update_fields.append('position')
                        updated_positions += 1
                        self.stdout.write(f"   💼 Должность для {employee.fio}: {position}")
                    
                    # Обновляем офис
                    if office and office.upper() in dict(Employee.OFFICE_CHOICES):
                        if employee.office == 'remote' or not employee.office:
                            employee.office = office.upper()
                            update_fields.append('office')
                            updated_offices += 1
                            self.stdout.write(f"   🏢 Офис для {employee.fio}: {office.upper()}")
                    
                    if update_fields:
                        employee.save(update_fields=update_fields)
                        
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"❌ Ошибка в строке {row_num}: {str(e)}"))
                
        self.stdout.write(self.style.SUCCESS(f"📊 Обновлено должностей: {updated_positions}"))
        self.stdout.write(self.style.SUCCESS(f"📊 Обновлено офисов: {updated_offices}"))