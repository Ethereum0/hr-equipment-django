import os
import logging
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee, Equipment

logger = logging.getLogger('import_equipment')

class Command(BaseCommand):
    help = 'Импорт оборудования из вкладки Выдано'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)

        if 'Выдано' in wb.sheetnames:
            ws = wb['Выдано']
            self.parse_equipment(ws)
        else:
            self.stdout.write(self.style.WARNING("Лист 'Выдано' не найден в файле."))

        self.stdout.write(self.style.SUCCESS('✅ Импорт Выдано завершён!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def find_employee_by_fio(self, fio_from_excel):
        """Ищет сотрудника по ФИО из Excel"""
        if not fio_from_excel:
            return None
            
        fio_clean = fio_from_excel.strip()
        
        # Пробуем найти по полному ФИО
        employee = Employee.objects.filter(fio__icontains=fio_clean).first()
        if employee:
            return employee
            
        # Если не нашли, пробуем разбить ФИО на части
        parts = fio_clean.split()
        if len(parts) >= 2:
            last_name = parts[0]
            first_name = parts[1]
            employee = Employee.objects.filter(
                last_name__icontains=last_name,
                first_name__icontains=first_name
            ).first()
            
        return employee

    def parse_equipment(self, ws):
        # Заголовки в СТРОКЕ 1 (как показала проверка)
        headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers_row)]

        col_map = {}
        for i, h in enumerate(headers):
            if 'ФИО сотрудника' in h:
                col_map['fio'] = i
            elif 'должность' in h:
                col_map['position'] = i
            elif 'офис' in h:
                col_map['office'] = i
            elif 'Перечень ТМЦ' in h:
                col_map['type'] = i
            elif 'Модель' in h:
                col_map['model'] = i
            elif 'Серийный номер' in h:
                col_map['serial'] = i
            elif 'MAC адрес' in h:
                col_map['mac'] = i
            elif 'Коментарий (IP/AnyDesk)' in h:
                col_map['ip_anydesk'] = i
            elif 'Коментарий' in h and i not in col_map.values():
                col_map['comment'] = i
            elif 'Ноут в домене' in h:
                col_map['domain_note'] = i

        processed = 0
        created = 0
        
        # Данные начинаются со СТРОКИ 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_list = [self.clean_cell(cell) for cell in row]
            
            # Пропускаем пустые строки
            if not any(row_list):
                continue
                
            fio = row_list[col_map.get('fio', -1)] if col_map.get('fio') is not None else ''
            if not fio:
                continue
                
            # Находим сотрудника
            employee = self.find_employee_by_fio(fio)
            if not employee:
                self.stdout.write(self.style.WARNING(f"⚠️ Строка {row_num}: Сотрудник '{fio}' не найден"))
                continue

            # Получаем данные оборудования
            type_name = row_list[col_map.get('type', -1)] if col_map.get('type') is not None else ''
            model = row_list[col_map.get('model', -1)] if col_map.get('model') is not None else ''
            serial = row_list[col_map.get('serial', -1)] if col_map.get('serial') is not None else ''
            mac = row_list[col_map.get('mac', -1)] if col_map.get('mac') is not None else ''
            ip_anydesk = row_list[col_map.get('ip_anydesk', -1)] if col_map.get('ip_anydesk') is not None else ''
            comment = row_list[col_map.get('comment', -1)] if col_map.get('comment') is not None else ''
            position = row_list[col_map.get('position', -1)] if col_map.get('position') is not None else ''
            office = row_list[col_map.get('office', -1)] if col_map.get('office') is not None else 'remote'
            
            # Обновляем должность сотрудника, если она указана
            if position and not employee.position:
                employee.position = position
                employee.save()
                
            # Создаем/обновляем оборудование
            if type_name and serial:
                equipment, created_flag = Equipment.objects.update_or_create(
                    serial_number=serial,
                    defaults={
                        'employee': employee,
                        'type': type_name,
                        'model': model,
                        'mac_address': mac,
                        'ip_or_anydesk': ip_anydesk,
                        'comment': comment,
                        'office': office.upper() if office else 'REMOTE',
                    }
                )
                
                if created_flag:
                    created += 1
                    self.stdout.write(self.style.SUCCESS(f"➕ Оборудование: {type_name} ({serial}) для {fio}"))
                    
            processed += 1
            
        self.stdout.write(self.style.SUCCESS(f"📊 Обработано: {processed} строк, создано: {created} оборудования"))