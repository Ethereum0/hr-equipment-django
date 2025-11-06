import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee, Equipment

class Command(BaseCommand):
    help = 'Импорт свободного оборудования из вкладки Свободное'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)

        if 'Свободное' in wb.sheetnames:
            ws = wb['Свободное']
            self.parse_free_equipment(ws)

        self.stdout.write(self.style.SUCCESS('✅ Импорт Свободного оборудования завершён!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def parse_free_equipment(self, ws):
        self.stdout.write("📊 Начало парсинга листа 'Свободное'")
        
        # Заголовки в строке 1
        headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers_row)]

        col_map = {}
        for i, h in enumerate(headers):
            if 'Перечень ТМЦ' in h:
                col_map['type'] = i
            elif 'Модель' in h:
                col_map['model'] = i
            elif 'Серийный номер' in h:
                col_map['serial'] = i
            elif 'MAC адрес' in h:
                col_map['mac'] = i
            elif 'Коментарий (IP/AnyDesk)' in h:
                col_map['ip_anydesk'] = i
            elif 'Коментарий' in h and i not in col_map.values():
                col_map['comment'] = i
            elif 'офис' in h.lower():
                col_map['office'] = i

        created_count = 0
        updated_count = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Пропускаем пустые строки
                continue
                
            type_name = self.clean_cell(row[col_map.get('type', -1)]) if col_map.get('type') is not None else ''
            model = self.clean_cell(row[col_map.get('model', -1)]) if col_map.get('model') is not None else ''
            serial = self.clean_cell(row[col_map.get('serial', -1)]) if col_map.get('serial') is not None else ''
            mac = self.clean_cell(row[col_map.get('mac', -1)]) if col_map.get('mac') is not None else ''
            ip_anydesk = self.clean_cell(row[col_map.get('ip_anydesk', -1)]) if col_map.get('ip_anydesk') is not None else ''
            comment = self.clean_cell(row[col_map.get('comment', -1)]) if col_map.get('comment') is not None else ''
            office = self.clean_cell(row[col_map.get('office', -1)]) if col_map.get('office') is not None else 'remote'
            
            # Нормализация офиса
            office_normalized = office.upper() if office else 'REMOTE'
            valid_offices = [choice[0] for choice in Employee.OFFICE_CHOICES]
            if office_normalized not in valid_offices:
                office_normalized = 'REMOTE'
            
            if type_name and serial:
                # Проверяем, не списано ли оборудование
                disposed = 'списано' in comment.lower() or 'списание' in comment.lower()
                
                equipment, created = Equipment.objects.update_or_create(
                    serial_number=serial,
                    defaults={
                        'employee': None,  # Всегда свободное
                        'type': type_name,
                        'model': model,
                        'mac_address': mac,
                        'ip_or_anydesk': ip_anydesk,
                        'comment': comment,
                        'office': office_normalized,
                        'disposed': disposed
                    }
                )
                
                if created:
                    created_count += 1
                    status = "📦 Свободное" if not disposed else "🗑️ Списано"
                    self.stdout.write(self.style.SUCCESS(f"{status}: {type_name} ({serial})"))
                else:
                    updated_count += 1
                    
        self.stdout.write(self.style.SUCCESS(f"📊 Итоги: создано {created_count}, обновлено {updated_count} свободного оборудования"))