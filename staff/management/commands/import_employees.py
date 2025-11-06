import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee

class Command(BaseCommand):
    help = 'Импорт сотрудников из вкладки Employment'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)
        
        if 'Employment' in wb.sheetnames:
            ws = wb['Employment']
            self.parse_employees(ws)
        
        self.stdout.write(self.style.SUCCESS('✅ Импорт сотрудников завершён!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def parse_employees(self, ws):
        # Заголовки в строке 1
        headers = ['num', 'first_last', 'middle', 'ad_login', 'pass_field']
        
        created_count = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[1]:  # Пропускаем пустые строки (колонка ФИ)
                continue
                
            first_last = self.clean_cell(row[1])  # ФИ (Фамилия Имя)
            middle = self.clean_cell(row[2])      # Отчество
            ad_login = self.clean_cell(row[3])    # AD логин
            
            # Формируем полное ФИО
            if first_last and middle:
                fio = f"{first_last} {middle}"
            else:
                fio = first_last
                
            if not fio or fio.strip() == '':
                continue
                
            # Создаем сотрудника
            employee, created = Employee.objects.get_or_create(
                ad_login=ad_login if ad_login else None,
                defaults={
                    'fio': fio.strip(),
                    'first_name': '',
                    'last_name': '', 
                    'middle_name': middle,
                    'status': 'active',
                    'office': 'remote'
                }
            )
            
            if created:
                created_count += 1
                self.stdout.write(self.style.SUCCESS(f"✅ Создан: {fio} (AD: {ad_login})"))
            else:
                self.stdout.write(self.style.WARNING(f"⚠️ Уже существует: {fio}"))
                
        self.stdout.write(self.style.SUCCESS(f"📊 Создано сотрудников: {created_count}"))