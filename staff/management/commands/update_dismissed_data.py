# staff/management/commands/update_dismissed_data.py
import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee

class Command(BaseCommand):
    help = 'Обновление данных существующих уволенных сотрудников'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)

        if 'Dismissal' in wb.sheetnames:
            ws = wb['Dismissal']
            self.update_dismissed_data(ws)

        self.stdout.write(self.style.SUCCESS('✅ Обновление данных уволенных завершено!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def update_dismissed_data(self, ws):
        self.stdout.write("📊 Обновление данных существующих уволенных")
        
        updated_count = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[1]:
                    continue
                    
                # Получаем все данные (как в предыдущем скрипте)
                first_last = self.clean_cell(row[1])
                middle = self.clean_cell(row[2])
                ad_login = self.clean_cell(row[3])
                ad_password = self.clean_cell(row[4])
                email = self.clean_cell(row[5])
                email_password = self.clean_cell(row[6])
                sip_number = self.clean_cell(row[7])
                web_3cx = self.clean_cell(row[8])
                pass_3cx = self.clean_cell(row[9])
                b24_login = self.clean_cell(row[10])
                b24_password = self.clean_cell(row[11])
                phone = self.clean_cell(row[12])
                tel_b24_login = self.clean_cell(row[15])
                tel_b24_password = self.clean_cell(row[16])
                sbis_login = self.clean_cell(row[17])
                sbis_password = self.clean_cell(row[18])
                
                # Формируем ФИО (ИСПРАВЛЕННЫЕ КАВЫЧКИ)
                if first_last and middle:
                    fio = f"{first_last} {middle}"
                else:
                    fio = first_last
                    
                if not fio:
                    continue
                
                # Ищем СУЩЕСТВУЮЩЕГО сотрудника
                employee = None
                if ad_login:
                    employee = Employee.objects.filter(ad_login=ad_login).first()
                if not employee:
                    employee = Employee.objects.filter(fio__iexact=fio).first()
                
                if employee:
                    # Обновляем только данные (не статус)
                    update_fields = []
                    
                    fields_to_update = {
                        'ad_password': ad_password,
                        'email': email,
                        'email_password': email_password,
                        'sip_number': str(sip_number) if sip_number else '',
                        'web_3cx': web_3cx,
                        'pass_3cx': pass_3cx,
                        'b24_login': b24_login,
                        'b24_password': b24_password,
                        'phone': phone,
                        'tel_b24_login': tel_b24_login,
                        'tel_b24_password': tel_b24_password,
                        'vats_login': sbis_login,
                        'vats_password': sbis_password,
                    }
                    
                    for field, value in fields_to_update.items():
                        current_value = getattr(employee, field)
                        if current_value != value:
                            setattr(employee, field, value)
                            update_fields.append(field)
                    
                    if update_fields:
                        employee.save()
                        updated_count += 1
                        self.stdout.write(self.style.SUCCESS(f"🔄 Обновлен: {employee.fio}"))
                        if update_fields:
                            self.stdout.write(f"   📝 Обновлены поля: {', '.join(update_fields)}")
                    
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"❌ Ошибка в строке {row_num}: {str(e)}"))
                
        self.stdout.write(self.style.SUCCESS(f"📊 Обновлено сотрудников: {updated_count}"))