# staff/management/commands/import_employees_final.py
import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee

class Command(BaseCommand):
    help = 'Полный импорт сотрудников со всеми полями из Employment'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)
        
        if 'Employment' in wb.sheetnames:
            ws = wb['Employment']
            self.parse_employees(ws)
        
        self.stdout.write(self.style.SUCCESS('✅ Импорт сотрудников завершён!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def parse_employees(self, ws):
        created_count = 0
        updated_count = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[1]:  # Пропускаем пустые строки (колонка ФИ)
                    continue
                    
                # Получаем данные из Excel
                first_last = self.clean_cell(row[1])  # ФИ
                middle = self.clean_cell(row[2])      # О (Отчество)
                ad_login = self.clean_cell(row[3])    # AD логин
                ad_password = self.clean_cell(row[4]) # Pass (AD пароль)
                email = self.clean_cell(row[5])       # Email
                email_password = self.clean_cell(row[6]) # Pass (Email пароль)
                sip_number = self.clean_cell(row[7])  # Номер SIP
                web_3cx = self.clean_cell(row[8])     # 3CX WEB пароль
                pass_3cx = self.clean_cell(row[9])    # Pass 3CX
                b24_login = self.clean_cell(row[10])  # B24
                b24_password = self.clean_cell(row[11]) # B24 Pass
                phone = self.clean_cell(row[12])      # телефон
                info = self.clean_cell(row[13])       # инфо
                manager = self.clean_cell(row[14])    # Рук (Руководитель)
                tel_b24_login = self.clean_cell(row[15])  # тел Б24 логин
                tel_b24_password = self.clean_cell(row[16])  # тел Б24 пароль
                sbis_login = self.clean_cell(row[17])  # СБИС логин
                sbis_password = self.clean_cell(row[18])  # СБИС пароль
                
                # Формируем полное ФИО
                if first_last and middle:
                    fio = f"{first_last} {middle}"
                else:
                    fio = first_last
                    
                if not fio:
                    continue
                
                # Разбиваем ФИО на компоненты
                parts = fio.split()
                last_name = parts[0] if len(parts) > 0 else ''
                first_name = parts[1] if len(parts) > 1 else ''
                
                # Обрабатываем email (если это формула CONCATENATE)
                if email and '=CONCATENATE' in email:
                    # Извлекаем email из формулы
                    if 'p-el.ru' in email and ad_login:
                        email = f"{ad_login}@p-el.ru"
                
                # Определяем офис по умолчанию
                office = 'remote'
                
                # Ищем существующего сотрудника
                employee = None
                if ad_login:
                    employee = Employee.objects.filter(ad_login=ad_login).first()
                
                if not employee and fio:
                    employee = Employee.objects.filter(fio__iexact=fio).first()
                
                # Подготовка данных для сохранения
                defaults = {
                    'fio': fio,
                    'first_name': first_name,
                    'last_name': last_name,
                    'middle_name': middle,
                    'ad_login': ad_login if ad_login else None,
                    'ad_password': ad_password if ad_password else '',
                    'email': email if email else '',
                    'email_password': email_password if email_password else '',
                    'sip_number': str(sip_number) if sip_number else '',
                    'web_3cx': web_3cx if web_3cx else '',
                    'pass_3cx': pass_3cx if pass_3cx else '',
                    'b24_login': b24_login if b24_login else '',
                    'b24_password': b24_password if b24_password else '',
                    'phone': phone if phone else '',
                    'info': info if info else '',
                    # 'supervisor': manager, # Это ForeignKey, нужно обработать отдельно
                    'tel_b24_login': tel_b24_login if tel_b24_login else '',
                    'tel_b24_password': tel_b24_password if tel_b24_password else '',
                    'vats_login': sbis_login if sbis_login else '',  # СБИС логин -> vats_login
                    'vats_password': sbis_password if sbis_password else '',  # СБИС пароль -> vats_password
                    'office': office,
                    'status': 'active',
                }
                
                if employee:
                    # Обновляем существующего сотрудника
                    update_fields = []
                    for field, value in defaults.items():
                        current_value = getattr(employee, field)
                        if current_value != value:
                            setattr(employee, field, value)
                            update_fields.append(field)
                    
                    if update_fields:
                        employee.save()
                        updated_count += 1
                        self.stdout.write(self.style.SUCCESS(f"🔄 Обновлен: {fio}"))
                        if update_fields:
                            self.stdout.write(f"   📝 Обновлены поля: {', '.join(update_fields)}")
                    else:
                        self.stdout.write(self.style.WARNING(f"⚠️ Без изменений: {fio}"))
                else:
                    # Создаем нового сотрудника
                    Employee.objects.create(**defaults)
                    created_count += 1
                    self.stdout.write(self.style.SUCCESS(f"✅ Создан: {fio}"))
                    self.stdout.write(f"   📧 Email: {email}")
                    self.stdout.write(f"   📞 SIP: {sip_number}")
                    
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"❌ Ошибка в строке {row_num}: {str(e)}"))
                import traceback
                self.stdout.write(self.style.ERROR(f"   Детали: {traceback.format_exc()}"))
                
        self.stdout.write(self.style.SUCCESS(f"📊 Итоги: создано {created_count}, обновлено {updated_count}"))