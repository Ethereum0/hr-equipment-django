# staff/management/commands/import_free_equipment_fixed.py
import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from staff.models import Employee, Equipment

class Command(BaseCommand):
    help = 'Импорт свободного оборудования из вкладки Свободное'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Путь к Excel-файлу', default='ТМЦ макет.xlsx')

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)

        if 'Свободное' in wb.sheetnames:
            ws = wb['Свободное']
            self.parse_free_equipment(ws)

        self.stdout.write(self.style.SUCCESS('✅ Импорт Свободного оборудования завершён!'))

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def parse_free_equipment(self, ws):
        self.stdout.write("📊 Начало парсинга листа 'Свободное'")
        
        created_count = 0
        updated_count = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):  # Пропускаем пустые строки
                    continue
                    
                type_name = self.clean_cell(row[3]) if len(row) > 3 else ''  # Перечень ТМЦ
                model = self.clean_cell(row[4]) if len(row) > 4 else ''      # Модель
                serial = self.clean_cell(row[5]) if len(row) > 5 else ''     # Серийный номер
                mac = self.clean_cell(row[6]) if len(row) > 6 else ''        # MAC адрес
                ip_anydesk = self.clean_cell(row[7]) if len(row) > 7 else '' # Коментарий (IP/AnyDesk)
                comment = self.clean_cell(row[8]) if len(row) > 8 else ''    # Коментарий
                office = self.clean_cell(row[2]) if len(row) > 2 else 'remote' # офис
                
                # Нормализация офиса
                office_normalized = office.upper() if office else 'REMOTE'
                valid_offices = [choice[0] for choice in Employee.OFFICE_CHOICES]
                if office_normalized not in valid_offices:
                    office_normalized = 'REMOTE'
                
                if type_name and serial:
                    # Проверяем, не списано ли оборудование
                    disposed = 'списано' in comment.lower() or 'списание' in comment.lower()
                    
                    equipment, created = Equipment.objects.update_or_create(
                        serial_number=serial,
                        defaults={
                            'employee': None,  # Всегда свободное
                            'type': type_name,
                            'model': model,
                            'mac_address': mac,
                            'ip_or_anydesk': ip_anydesk,
                            'comment': comment,
                            'office': office_normalized,
                            'disposed': disposed
                        }
                    )
                    
                    if created:
                        created_count += 1
                        status = "📦 Свободное" if not disposed else "🗑️ Списано"
                        self.stdout.write(self.style.SUCCESS(f"{status}: {type_name} ({serial})"))
                    else:
                        updated_count += 1
                        
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"❌ Ошибка в строке {row_num}: {str(e)}"))
                    
        self.stdout.write(self.style.SUCCESS(f"📊 Итоги: создано {created_count}, обновлено {updated_count} свободного оборудования"))